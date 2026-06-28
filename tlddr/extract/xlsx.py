from pathlib import Path
from openpyxl import load_workbook
from tlddr.extract.base import ExtractContext
from tlddr.ids import doc_id, sha256_file
from tlddr.models import ExtractedDoc, PageProvenance, SignalType, ExtractMethod

MAX_ROWS = 200  # cap dump per sheet for the reconnaissance report


def _cell(value: object) -> str:
    return "" if value is None else str(value)


def extract(path: Path, ctx: ExtractContext) -> ExtractedDoc:
    wb = load_workbook(path, data_only=True, read_only=False)
    parts: list[str] = []
    pages: list[PageProvenance] = []
    warnings: list[str] = []

    try:
        for index, ws in enumerate(wb.worksheets, start=1):
            parts.append(f"--- sheet: {ws.title} ---")
            rows = 0
            for row in ws.iter_rows(values_only=True):
                if rows >= MAX_ROWS:
                    warnings.append(f"sheet '{ws.title}' truncated at {MAX_ROWS} rows")
                    break
                line = " | ".join(_cell(v) for v in row)
                if line.strip(" |"):
                    parts.append(line)
                    rows += 1
            if ws.merged_cells.ranges:
                warnings.append(f"messy: sheet '{ws.title}' has merged cells")
            pages.append(PageProvenance(
                page=index, method=ExtractMethod.OPENPYXL_XLSX,
                has_text_layer=True, char_count=0,
            ))
    finally:
        wb.close()

    return ExtractedDoc(
        id=doc_id(path),
        source_path=str(path),
        source_sha256=sha256_file(path),
        signal_type=SignalType.SPREADSHEET,
        raw_title=path.stem,
        content="\n".join(parts),
        pages=pages,
        warnings=warnings,
        extractor="xlsx",
    )
